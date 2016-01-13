import pdb
import urllib2
import re
from openpyxl import load_workbook
from openpyxl import Workbook

urls = [u'http://новостройки-в-волгограде.рф/%D0%BD%D0%BE%D0%B2%D0%BE%D1%81%D1%82%D1%80%D0%BE%D0%B9%D0%BA%D0%B8-%D1%81%D0%BE%D0%B2%D0%B5%D1%82%D1%81%D0%BA%D0%BE%D0%B3%D0%BE/item/3-%D0%B6%D0%BA-%D1%80%D0%BE%D0%B4%D0%BD%D0%B8%D0%BA%D0%BE%D0%B2%D0%B0%D1%8F-%D0%B4%D0%BE%D0%BB%D0%B8%D0%BD%D0%B0.html',
        u'http://новостройки-в-волгограде.рф/%D0%BD%D0%BE%D0%B2%D0%BE%D1%81%D1%82%D1%80%D0%BE%D0%B9%D0%BA%D0%B8-%D1%81%D0%BE%D0%B2%D0%B5%D1%82%D1%81%D0%BA%D0%BE%D0%B3%D0%BE/item/11-%D0%B6%D0%BA-%D0%BD%D0%BE%D0%B2%D1%8B%D0%B9-%D1%81%D0%B2%D0%B5%D1%82.html',
        u'http://новостройки-в-волгограде.рф/%D0%BD%D0%BE%D0%B2%D0%BE%D1%81%D1%82%D1%80%D0%BE%D0%B9%D0%BA%D0%B8-%D1%81%D0%BE%D0%B2%D0%B5%D1%82%D1%81%D0%BA%D0%BE%D0%B3%D0%BE/item/249-%D0%B6%D0%BA-%D0%BA%D0%BE%D0%BC%D0%B0%D1%80%D0%BE%D0%B2%D0%BE.html',
        u'http://новостройки-в-волгограде.рф/%D0%BD%D0%BE%D0%B2%D0%BE%D1%81%D1%82%D1%80%D0%BE%D0%B9%D0%BA%D0%B8-%D1%81%D0%BE%D0%B2%D0%B5%D1%82%D1%81%D0%BA%D0%BE%D0%B3%D0%BE/item/14-%D0%B6%D0%BA-%D0%B4%D0%BE%D0%BC%D0%B8%D0%BD%D0%B0%D0%BD%D1%82.html',
        u'http://новостройки-в-волгограде.рф/%D0%BD%D0%BE%D0%B2%D0%BE%D1%81%D1%82%D1%80%D0%BE%D0%B9%D0%BA%D0%B8-%D1%81%D0%BE%D0%B2%D0%B5%D1%82%D1%81%D0%BA%D0%BE%D0%B3%D0%BE/item/9-%D0%B6%D0%BA-%D0%BA%D0%BE%D0%BC%D0%B0%D1%80%D0%BE%D0%B2%D0%BE.html',
        u'http://новостройки-в-волгограде.рф/%D0%BD%D0%BE%D0%B2%D0%BE%D1%81%D1%82%D1%80%D0%BE%D0%B9%D0%BA%D0%B8-%D1%81%D0%BE%D0%B2%D0%B5%D1%82%D1%81%D0%BA%D0%BE%D0%B3%D0%BE/item/263-%D0%BA%D1%83%D0%BF%D0%B8%D1%82%D1%8C-%D0%BA%D0%B2%D0%B0%D1%80%D1%82%D0%B8%D1%80%D1%83-%D0%B2-%D0%B6%D0%B8%D0%BB%D0%BE%D0%BC-%D0%BA%D0%BE%D0%BC%D0%BF%D0%BB%D0%B5%D0%BA%D1%81%D0%B5-%D0%B0%D0%B2%D1%80%D0%BE%D1%80%D0%B0-%D0%B0%D0%B2%D0%B0%D0%BD%D0%B3%D0%B0%D1%80%D0%B4.html',
        u'http://новостройки-в-волгограде.рф/%D0%BD%D0%BE%D0%B2%D0%BE%D1%81%D1%82%D1%80%D0%BE%D0%B9%D0%BA%D0%B8-%D1%81%D0%BE%D0%B2%D0%B5%D1%82%D1%81%D0%BA%D0%BE%D0%B3%D0%BE/item/304-%D0%BA%D0%BB%D1%83%D0%B1%D0%BD%D1%8B%D0%B9-%D0%B6%D0%B8%D0%BB%D0%BE%D0%B9-%D0%BA%D0%BE%D0%BC%D0%BF%D0%BB%D0%B5%D0%BA%D1%81-%D0%BA%D0%BE%D0%BA%D0%BE%D1%81.html',
        u'http://новостройки-в-волгограде.рф/%D0%BD%D0%BE%D0%B2%D0%BE%D1%81%D1%82%D1%80%D0%BE%D0%B9%D0%BA%D0%B8-%D1%81%D0%BE%D0%B2%D0%B5%D1%82%D1%81%D0%BA%D0%BE%D0%B3%D0%BE/item/13-%D0%B6%D0%BA-%D0%BD%D0%BE%D0%B2%D1%8B%D0%B9-%D1%81%D0%B2%D0%B5%D1%82.html',
        u'http://новостройки-в-волгограде.рф/%D0%BD%D0%BE%D0%B2%D0%BE%D1%81%D1%82%D1%80%D0%BE%D0%B9%D0%BA%D0%B8-%D1%81%D0%BE%D0%B2%D0%B5%D1%82%D1%81%D0%BA%D0%BE%D0%B3%D0%BE/item/12-%D0%B6%D0%B4-%D0%BA%D0%BE%D0%BC%D0%BF%D0%BE%D0%B7%D0%B8%D1%86%D0%B8%D1%8F.html',
        u'http://новостройки-в-волгограде.рф/%D0%BD%D0%BE%D0%B2%D0%BE%D1%81%D1%82%D1%80%D0%BE%D0%B9%D0%BA%D0%B8-%D1%81%D0%BE%D0%B2%D0%B5%D1%82%D1%81%D0%BA%D0%BE%D0%B3%D0%BE/item/10-%D0%B6%D0%B4-%D0%B4%D0%BE%D0%BC-%D0%BD%D0%B0-%D0%BF%D0%B5%D1%82%D1%80%D0%BE%D0%B2%D1%81%D0%BA%D0%BE%D0%B9.html',
        u'http://новостройки-в-волгограде.рф/%D0%BD%D0%BE%D0%B2%D0%BE%D1%81%D1%82%D1%80%D0%BE%D0%B9%D0%BA%D0%B8-%D0%BA%D1%80%D0%B0%D1%81%D0%BD%D0%BE%D0%BE%D0%BA%D1%82%D1%8F%D0%B1%D1%80%D1%8C%D1%81%D0%BA%D0%BE%D0%B3%D0%BE/item/25-%D0%B6%D0%B4-%D0%BC%D0%B8%D1%88%D0%B8%D0%BD%D0%BE.html',
        u'http://новостройки-в-волгограде.рф/%D0%BD%D0%BE%D0%B2%D0%BE%D1%81%D1%82%D1%80%D0%BE%D0%B9%D0%BA%D0%B8-%D0%BA%D1%80%D0%B0%D1%81%D0%BD%D0%BE%D0%BE%D0%BA%D1%82%D1%8F%D0%B1%D1%80%D1%8C%D1%81%D0%BA%D0%BE%D0%B3%D0%BE/item/21-%D0%B6%D0%B4-%D1%81%D0%BE%D0%BB%D0%BD%D0%B5%D1%87%D0%BD%D1%8B%D0%B9-%D0%B3%D0%BE%D1%80%D0%BE%D0%B4.html',
        u'http://новостройки-в-волгограде.рф/%D0%BD%D0%BE%D0%B2%D0%BE%D1%81%D1%82%D1%80%D0%BE%D0%B9%D0%BA%D0%B8-%D0%BA%D1%80%D0%B0%D1%81%D0%BD%D0%BE%D0%BE%D0%BA%D1%82%D1%8F%D0%B1%D1%80%D1%8C%D1%81%D0%BA%D0%BE%D0%B3%D0%BE/item/26-%D0%B6%D0%BA-%D0%BE%D0%BB%D0%B5%D0%B6%D0%BA%D0%B8%D0%BD%D0%B0-%D1%81%D0%BB%D0%BE%D0%B1%D0%BE%D0%B4%D0%B0.html',
        u'http://новостройки-в-волгограде.рф/%D0%BD%D0%BE%D0%B2%D0%BE%D1%81%D1%82%D1%80%D0%BE%D0%B9%D0%BA%D0%B8-%D0%BA%D0%B8%D1%80%D0%BE%D0%B2%D1%81%D0%BA%D0%BE%D0%B3%D0%BE/item/17-%D0%B6%D0%BA-%D1%81%D0%B0%D0%BD%D0%B0%D1%82%D0%BE%D1%80%D0%BD%D1%8B%D0%B9.html',
        u'http://новостройки-в-волгограде.рф/%D0%BD%D0%BE%D0%B2%D0%BE%D1%81%D1%82%D1%80%D0%BE%D0%B9%D0%BA%D0%B8-%D0%BA%D0%B8%D1%80%D0%BE%D0%B2%D1%81%D0%BA%D0%BE%D0%B3%D0%BE/item/16-%D0%B6%D0%BA-%D0%BF%D0%B0%D1%80%D0%BA-%D0%B5%D0%B2%D1%80%D0%BE%D0%BF%D0%B5%D0%B9%D1%81%D0%BA%D0%B8%D0%B92.html',
        u'http://новостройки-в-волгограде.рф/%D0%BD%D0%BE%D0%B2%D0%BE%D1%81%D1%82%D1%80%D0%BE%D0%B9%D0%BA%D0%B8-%D0%BA%D0%B8%D1%80%D0%BE%D0%B2%D1%81%D0%BA%D0%BE%D0%B3%D0%BE/item/15-%D0%B6%D0%BA-%D0%BF%D0%B0%D1%80%D0%BA-%D0%B5%D0%B2%D1%80%D0%BE%D0%BF%D0%B5%D0%B9%D1%81%D0%BA%D0%B8%D0%B91.html',
        u'http://новостройки-в-волгограде.рф/%D0%BD%D0%BE%D0%B2%D0%BE%D1%81%D1%82%D1%80%D0%BE%D0%B9%D0%BA%D0%B8-%D0%BA%D0%B8%D1%80%D0%BE%D0%B2%D1%81%D0%BA%D0%BE%D0%B3%D0%BE/item/416-%D0%B6%D0%B4-%D0%BA%D0%B8%D1%80%D0%BE%D0%B2%D0%B0-101.html',
        u'http://новостройки-в-волгограде.рф/%D0%BD%D0%BE%D0%B2%D0%BE%D1%81%D1%82%D1%80%D0%BE%D0%B9%D0%BA%D0%B8-%D0%BA%D0%B8%D1%80%D0%BE%D0%B2%D1%81%D0%BA%D0%BE%D0%B3%D0%BE/item/22-%D0%B6%D0%B4-503-%D0%BC%D0%BA%D1%80.html',
        u'http://новостройки-в-волгограде.рф/%D0%BD%D0%BE%D0%B2%D0%BE%D1%81%D1%82%D1%80%D0%BE%D0%B9%D0%BA%D0%B8-%D1%82%D1%80%D0%B0%D0%BA%D1%82%D0%BE%D1%80%D0%BE%D0%B7%D0%B0%D0%B2%D0%BE%D0%B4%D1%81%D0%BA%D0%BE%D0%B3%D0%BE/item/19-%D0%B6%D0%B4-%D1%81%D0%B5%D0%B2%D0%B5%D1%80%D0%BD%D0%BE%D0%B5-%D1%81%D0%B8%D1%8F%D0%BD%D0%B8%D0%B5.html',
        u'http://новостройки-в-волгограде.рф/%D0%BD%D0%BE%D0%B2%D0%BE%D1%81%D1%82%D1%80%D0%BE%D0%B9%D0%BA%D0%B8-%D1%82%D1%80%D0%B0%D0%BA%D1%82%D0%BE%D1%80%D0%BE%D0%B7%D0%B0%D0%B2%D0%BE%D0%B4%D1%81%D0%BA%D0%BE%D0%B3%D0%BE/item/18-%D0%B6%D0%B4-%D0%BA%D0%BE%D0%BB%D1%83%D0%BC%D0%B1%D0%B0-7.html',
        u'http://новостройки-в-волгограде.рф/%D0%BD%D0%BE%D0%B2%D0%BE%D1%81%D1%82%D1%80%D0%BE%D0%B9%D0%BA%D0%B8-%D1%82%D1%80%D0%B0%D0%BA%D1%82%D0%BE%D1%80%D0%BE%D0%B7%D0%B0%D0%B2%D0%BE%D0%B4%D1%81%D0%BA%D0%BE%D0%B3%D0%BE/item/270-%D0%B6%D0%B4-%D0%B3%D0%B5%D1%80%D0%BE%D0%B5%D0%B2-%D1%82%D1%83%D0%BB%D1%8B-7.html',
        u'http://новостройки-в-волгограде.рф/%D0%BD%D0%BE%D0%B2%D0%BE%D1%81%D1%82%D1%80%D0%BE%D0%B9%D0%BA%D0%B8-%D0%BA%D1%80%D0%B0%D1%81%D0%BD%D0%BE%D0%B0%D1%80%D0%BC%D0%B5%D0%B9%D1%81%D0%BA%D0%BE%D0%B3%D0%BE/item/262-%D0%BA%D1%83%D0%BF%D0%B8%D1%82%D1%8C-%D0%BA%D0%B2%D0%B0%D1%80%D1%82%D0%B8%D1%80%D1%83-%D0%B2-%D0%B6%D0%B8%D0%BB%D0%BE%D0%BC-%D0%BA%D0%BE%D0%BC%D0%BF%D0%BB%D0%B5%D0%BA%D1%81%D0%B5-%D0%B0%D0%B4%D0%BC%D0%B8%D1%80%D0%B0%D0%BB%D1%82%D0%B5%D0%B9%D1%81%D0%BA%D0%B8%D0%B9-%D0%BD%D0%B0-%D1%84%D0%B0%D0%B4%D0%B5%D0%B5%D0%B2%D0%B0-53-%D0%B7%D0%B6%D0%B1%D0%B8%D0%BA.html',
        u'http://новостройки-в-волгограде.рф/%D0%BD%D0%BE%D0%B2%D0%BE%D1%81%D1%82%D1%80%D0%BE%D0%B9%D0%BA%D0%B8-%D0%BA%D1%80%D0%B0%D1%81%D0%BD%D0%BE%D0%B0%D1%80%D0%BC%D0%B5%D0%B9%D1%81%D0%BA%D0%BE%D0%B3%D0%BE/item/20-%D0%B6%D0%B4-%D1%84%D0%B0%D0%B4%D0%B5%D0%B5%D0%B2%D0%B0-63.html',
        u'http://новостройки-в-волгограде.рф/%D0%BD%D0%BE%D0%B2%D0%BE%D1%81%D1%82%D1%80%D0%BE%D0%B9%D0%BA%D0%B8-%D0%BA%D1%80%D0%B0%D1%81%D0%BD%D0%BE%D0%B0%D1%80%D0%BC%D0%B5%D0%B9%D1%81%D0%BA%D0%BE%D0%B3%D0%BE/item/215-%D0%BA%D1%83%D0%BF%D0%B8%D1%82%D1%8C-%D0%BA%D0%B2%D0%B0%D1%80%D1%82%D0%B8%D1%80%D1%83-%D0%B2-%D0%B6%D0%B8%D0%BB%D0%BE%D0%BC-%D0%BA%D0%BE%D0%BC%D0%BF%D0%BB%D0%B5%D0%BA%D1%81%D0%B5-50-%D0%BB%D0%B5%D1%82-%D0%BE%D0%BA%D1%82%D1%8F%D0%B1%D1%80%D1%8F-36.html']

result = []
wb = Workbook(optimized_write = True)
ws = wb.create_sheet()
wb.save(r'D:\TEMP\new.xlsx')

def parse_urls(base):
    index = 0
    for url in base:
        domain = url[7:34].encode('idna')
        finish = url[34:]
        try:
            fetch_url = urllib2.urlopen('http://'+domain+finish)
            strrr = fetch_url.readlines()
            fetch_url.close()
            print ('-'*50)
            print ('HTML code of URL =', url)
            print ('-'*50)
            result.append([url])
            result[index].extend(parse_lines(strrr))

            wb = load_workbook(r'D:\TEMP\new.xlsx')
            ws = wb.active
            row_index = ws.get_highest_row()
            for j in range(len(result[index])):
                ws.cell(row = row_index, column = j).value = result[index][j]
            wb.save(r'D:\TEMP\new.xlsx')

            index = index + 1            
        except IOError:
            print ('Cannot open URL %s for reading' % url)
            str1 = 'error!'
            return

   
def parse_lines(input_page):
    #pdb.set_trace()
    longt, latid, title, district, floor, price, end_term = '','','','','','',''
    for line_encode in input_page:
        line = line_encode.decode('utf-8')
        if u'myMap.geoObjects.add(new ymaps.Placemark([44' in line:
            longLat = line[line.find('[')+1:line.find(']')]
            longt = longLat[:line.find(',')]
            latid = longLat[line.find(',')+1:]
        elif u'<meta name="title"' in line:
            start = '<meta name="title" content="'
            end = '" />'
            title = re.search('%s(.*)%s' % (start, end), line).group(1)
        elif u'<strong><span>Район:</span></strong>' in line:
            start = '<span style="color: #ff6600;">'
            end = '</span>'
            district = re.search('%s(.*)%s' % (start, end), line).group(1)
        elif u'<strong><span style="color: #65a8e9;">Этажность: </span></strong>' in line:
            start = '<span style="color: #ff6600;">'
            end = '</span></span></p>'
            floor = re.search('%s(.*)%s' % (start, end), line).group(1)
        elif u'<span style="color: #65a8e9;"><strong>Цена за 1 кв.м.</strong>' in line:
            start = '<span style="color: #ff6600;">'
            end = '</span></p>'
            price = re.search('%s(.*)%s' % (start, end), line).group(1)
        elif u'<strong><span style="color: #65a8e9;">Срок сдачи:</span></strong>' in line:
            start = '<span style="color: #ff6600;">'
            end = '</span></p>'
            end_term = re.search('%s(.*)%s' % (start, end), line).group(1)

    result = [latid, longt, title, district, floor, price, end_term]
    return result

parse_urls(urls)
print result
