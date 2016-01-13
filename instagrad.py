from openpyxl import load_workbook
from openpyxl import Workbook
import pdb

def workout(input_path, text_prefix, filename):
    wb = load_workbook(filename = input_path)
    ws = wb.active
    
    wb_new = Workbook()
    ws_new = wb_new.active
    iter = 0
    
    text, start, stop = text_prefix, 0, ws.get_highest_row()
    reminder = ws.get_highest_row() % 8
    
    while start < stop - reminder:
        for i in range(8): # i - only iterates 8 times and gives index for ws_new
            if ws.cell(row=start, column=2).internal_value == None or type(ws.cell(row=start, column=2).internal_value) is int:
                ws.cell(row=start, column=2).value = ""
            if ws.cell(row=start, column=2).internal_value == None:
                ws.cell(row=start, column=2).value = ""
            text = text + ws.cell(row=start, column=1).internal_value + ": " + ws.cell(row=start, column=2).internal_value + " " + ws.cell(row=start, column=3).internal_value + "\n"
            ws_new.cell(row = iter, column = i + 2).value = ws.cell(row=start, column=4).internal_value
            start = start + 1
        all_text = ws_new.cell(row = iter, column = 0)
        all_text.value = text
        text, iter = text_prefix, iter + 1
    
    # pdb.set_trace()
    
    while start < stop:
        i = 0
        if ws.cell(row=start, column=2).internal_value == None:
            ws.cell(row=start, column=2).value = ""
        text = text + ws.cell(row=start, column=1).internal_value + ": " + ws.cell(row=start, column=2).internal_value + " " + ws.cell(row=start, column=3).internal_value + "\n"
        ws_new.cell(row = iter, column = i + 2).value = ws.cell(row=start, column=4).internal_value
        start, i = start + 1, i + 1
    
    all_text = ws_new.cell(row = iter, column = 0)
    all_text.value = text
    text, iter = text_prefix, iter + 1
    
    wb_new.save(filename)

workout(r'C:\Users\User\Downloads\New photo by anyone in 7Vetr.xlsx', u'7Ветров\n', r'C:\Users\User\Downloads\7vetrov.xlsx')
workout(r'C:\Users\User\Downloads\New photo by anyone in Vorosh.xlsx', u'Ворошиловский\n', r'C:\Users\User\Downloads\Vorosh.xlsx')
workout(r'C:\Users\User\Downloads\New photo by anyone in Tsentr.xlsx', u'Центр\n', r'C:\Users\User\Downloads\Tsentr.xlsx')
workout(r'C:\Users\User\Downloads\New photo by anyone in Tspkio.xlsx', u'ЦПКиО\n', r'C:\Users\User\Downloads\Stad.xlsx')
